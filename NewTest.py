import openpyxl, xlrd
from customtkinter import *
from ultralytics import YOLO
from tkinter import simpledialog
import tkinter
from tkinter import ttk, messagebox
import serial.tools.list_ports
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import serial.tools.list_ports
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
from scipy import signal
import datetime
import sqlite3
from datetime import datetime
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import ttk, messagebox
import cv2
import io
import numpy as np
import pyrealsense2 as rs
import os
import serial
import threading
import time
import queue


# Configuration
DATABASE_FILE = 'Rail_Inspection_database.db'
REPORT_FILE = 'defect_report.html'
IMAGE_OUTPUT_DIR = 'report_images'
class ImageDatabase:
    def __init__(self, db_name='Rail_Inspection_database.db'):
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self._create_table()

    def _create_table(self):
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            defect_class TEXT, 
            gps_coordinates TEXT,
            distance TEXT,
            gauge TEXT,
            cant TEXT,
            action_taken TEXT,
            image_data BLOB NOT NULL,
            thumbnail BLOB NOT NULL
        )
        ''')
        self.conn.commit()

    def save_image(self, image_data, defect_class="", gps_coordinates="", distance = "", gauge ="", cant = "", action_taken="None"):
        """Save image with metadata to database"""
        try:
            # Create thumbnail
            img = Image.open(io.BytesIO(image_data))
            img.thumbnail((200, 200))

            thumbnail_bytes = io.BytesIO()
            img.save(thumbnail_bytes, format='JPEG')

            self.cursor.execute('''
            INSERT INTO images (timestamp, defect_class, gps_coordinates, distance, gauge, cant, action_taken, image_data, thumbnail)
            VALUES (?, ?, ?, ?, ?, ?,?, ?, ?)
            ''', (
                datetime.now().isoformat(),  # Correct usage
                defect_class,
                gps_coordinates,
                distance,
                gauge,
                cant,
                action_taken,
                image_data,
                thumbnail_bytes.getvalue()
            ))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error saving image: {e}")
            return False

    def get_all_images(self):
        """Retrieve all image thumbnails with metadata"""
        self.cursor.execute('''
        SELECT id, timestamp, defect_class, gps_coordinates, distance, action_taken, thumbnail FROM images
        ORDER BY timestamp DESC
        ''')
        return self.cursor.fetchall()

    def get_full_image(self, image_id):
        """Retrieve full image by ID"""
        self.cursor.execute('''
        SELECT image_data FROM images WHERE id = ?
        ''', (image_id,))
        result = self.cursor.fetchone()
        return result[0] if result else None

    def close(self):
        self.conn.close()

class UTImageDatabase:
    def __init__(self, db_name='Ultrasonic_Rail_Inspection_database.db'):
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self._create_table()

    def _create_table(self):
        self.cursor.execute('''
        CREATE TABLE IF NOT EXISTS images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            defect_class TEXT, 
            gps_coordinates TEXT,
            image_data BLOB NOT NULL,
            thumbnail BLOB NOT NULL
        )
        ''')
        self.conn.commit()

    def save_image(self, image_data, defect_class="", gps_coordinates=""):
        """Save image with metadata to database"""
        try:
            # Create thumbnail
            img = Image.open(io.BytesIO(image_data))
            img.thumbnail((200, 200))

            thumbnail_bytes = io.BytesIO()
            img.save(thumbnail_bytes, format='JPEG')

            self.cursor.execute('''
            INSERT INTO images (timestamp, defect_class, gps_coordiantes, image_data, thumbnail)
            VALUES (?, ?, ?, ?, ?)
            ''', (
                datetime.datetime.now().isoformat(),
                defect_class,
                gps_coordinates,
                image_data,
                thumbnail_bytes.getvalue()
            ))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Error saving image: {e}")
            return False

    def get_all_images(self):
        """Retrieve all image thumbnails with metadata"""
        self.cursor.execute('''
        SELECT id, timestamp, defect_class, gps_coordinates, thumbnail FROM images
        ORDER BY timestamp DESC
        ''')
        return self.cursor.fetchall()

    def get_full_image(self, image_id):
        """Retrieve full image by ID"""
        self.cursor.execute('''
        SELECT image_data FROM images WHERE id = ?
        ''', (image_id,))
        result = self.cursor.fetchone()
        return result[0] if result else None

    def close(self):
        self.conn.close()

class YOLOv8_GUI(tk.Frame):
    def __init__(self, master):
        super().__init__(master)  # Initialize the root window
        self.master = master
        self.master.title("Railway Visual Track Inspection")
        self.master.configure(bg="#326273")

        # Maximize the window (works on Windows, Linux)
        self.master.state('zoomed')  # Alternative: self.root.attributes('-fullscreen', True)
        self.grid()

        # Initialize database
        self.db = ImageDatabase()

        # Initialize models
        self.seg_model1 = YOLO("Weights/Fasteners/best.pt")
        self.seg_model2 = YOLO("Weights/Rails/best.pt")
        self.obb_model = YOLO("Weights/Sleepers/best.pt")

        # Video capture
        self.cap = None
        self.video_path = ""
        self.is_running = False
        self.current_frame = None
        self.photo = None

        # Variables for serial communication
        self.defect_class = "___"
        self.gps_coordinates = "__"
        self.is_repairing = False
        self.gauge = "___"
        self.cant = "___"
        self.action = "___"

        self.gps_variable = ""
        self.cant_variable = ""
        self.gauge_variable = ""
        self.action_variable = ""
        self.class_variable = ""
        self.distance_variable = ""
        self.demo_variable = ""

        self.defective_rails = 0
        self.poor_sleepers =0
        self.fair_sleepers = 0
        self.good_sleepers = 0
        self.missing_dogscrews = 0
        self.defective_fasteners = 0
        self.defective_joints = 0

        #Move Variables
        self.move_scan = False
        self.move_repair = False
        self.move_cruse = False
        self.move_stop = False
        self.move_reverse = False

        # Serial Communication with Arduino
        self.serial_port = 'COM6'  # <--- IMPORTANT: Change to your Arduino's serial port (e.g., 'COM3' on Windows, '/dev/ttyUSB0' on Linux)
        self.baud_rate = 9600

        self.ser = None
        self.serial_thread = None
        self.running = False
        self.data_queue = queue.Queue()  # Thread-safe queue for data

        self.port = None
        self.data = []
        self.connected = False
        #self.ser = None

        # Configure grid weights for expansion
        self.setup_grid_configuration()

        # Create GUI
        self.create_widgets()

        # Bind window resize event
        self.master.bind('<Configure>', self.on_window_resize)

        if self.move_scan or self.move_repair:
            self.update_gps_cant_gauge()

    def on_window_resize(self, event):
        # This ensures proper resizing when window is restored/maximized
        self.master.update_idletasks()

    def setup_grid_configuration(self):
        # Make all rows and columns expandable
        for i in range(3):  # For 3 rows
            self.master.grid_rowconfigure(i, weight=1)
        for j in range(3):  # For 3 columns
            self.master.grid_columnconfigure(j, weight=1)


    def create_widgets(self):

        # Frame for video display
        self.video_frame = tk.LabelFrame(self.master, bg="#326273",font=('Arial', 14), text="Detection Output")
        self.video_frame.grid( row=0, column=0, padx=10, pady=10)

        self.canvas = tk.Canvas(self.video_frame, bg='black', width=840, height=500)
        self.canvas.grid(row=0, column=0, sticky="nws", padx=10, pady=10)

        # Control buttons frame
        self.control_frame = tk.Frame(self.master, bg="#326273", height = 20)
        self.control_frame.grid(row =1, column = 0, sticky = "news", padx =10, pady=10)

        # Video controls
        style = ttk.Style() #Button Style
        # Create a custom style
        style.configure('LargeFont.TButton',
                        font=('Arial', 11),
                        background="#326273",  # Main color
                        #foreground="white",  # Text color
                        bordercolor="#326273",  # Border color
                        darkcolor="#326273",  # Dark shadow
                        lightcolor="#326273",  # Light highlight
                        focuscolor="#326273",  # Focus ring
                        #relief="flat",  # Remove 3D effect
                        padding=5,
                        borderwidth=2,
                        width = 15,
                        height = 1,
                        relief='raised')
        #style.configure('LargeFont.TButton', font=('Arial', 11))

        btn_open = tk.Button(self.control_frame, text="Open Video", command=self.open_video,
                             bg = "#326273", foreground="white", width=15, height=1, font=('Arial', 11))
        btn_open.grid(row=0, column=0, padx=10)

        btn_start = tk.Button(self.control_frame, text="Start", command=self.start_video,
                               bg = "#326273", foreground="white", width=15, height=1, font=('Arial', 11))
        btn_start.grid(row=0, column=1, padx=10)

        btn_stop = tk.Button(self.control_frame, text="Stop", command=self.stop_video,
                             bg = "#326273", foreground="white", width=15, height=1, font=('Arial', 11))
        btn_stop.grid(row=0, column=2, padx=10)

        btn_snapshot = tk.Button(self.control_frame, text="Snapshot", command=self.take_snapshot,
                                 bg = "#326273", foreground="white", width=15, height=1, font=('Arial', 11))
        btn_snapshot.grid(row=0, column=3, padx=10)

        # Image detection
        btn_open_img = tk.Button(self.control_frame, text="Open Image", command=self.open_image,
                                 bg = "#326273", foreground="white", width=15, height=1, font=('Arial', 11))
        btn_open_img.grid(row=0, column=4, pady =10, padx=10)

        btn_detect_img = tk.Button(self.control_frame, text="Detect Image", command=self.detect_image,
                                   bg = "#326273", foreground="white", width=15, height=1, font=('Arial', 11))
        btn_detect_img.grid(row=0, column=5, pady = 10, padx=10)

        btn_livestream_vid = tk.Button(self.control_frame, text = "Livestream", command = self.livestream,
                                       bg = "#326273", foreground="white", width=15, height=1, font=('Arial', 11))
        btn_livestream_vid.grid(row=0, column=5, pady = 10, padx=10)

        # Model selection checkboxes
        model_frame = tk.LabelFrame(self.master, bg="#326273",
                                    text="Model Selection",
                                    font = ("Arial, 14"))
        model_frame.grid(row = 2, column = 0, sticky = "news", pady=10, padx=10)

        self.seg1_var = tk.BooleanVar(value=True)
        self.seg2_var = tk.BooleanVar(value=True)
        self.obb_var = tk.BooleanVar(value=True)
        self.autosave_var = tk.BooleanVar(value=False)

        chk_seg1 = tk.Checkbutton(model_frame, text="Segmentation Model 1", variable=self.seg1_var,
                                  bg="#326273", foreground="white", selectcolor="#326273", font=('Ariel', 13))
        chk_seg1.grid(row=0, column=0,pady = 10, padx=10)

        chk_seg2 = tk.Checkbutton(model_frame, text="Segmentation Model 2", variable=self.seg2_var,
                                  bg="#326273", foreground="white", selectcolor="#326273", font=('Ariel', 13))
        chk_seg2.grid(row=0, column=1, pady = 10, padx=10)#, sticky=tk.W)

        chk_obb = tk.Checkbutton(model_frame, text="OBB Model", variable=self.obb_var,
                                  bg="#326273", foreground="white", selectcolor="#326273", font=('Ariel', 13))
        chk_obb.grid(row=0, column=2,  pady = 10, padx=10)

        chk_autosave = tk.Checkbutton(model_frame, text="Autosave", variable=self.autosave_var,
                                      bg="#326273", foreground="white", selectcolor="#326273", font=('Ariel', 13))
        chk_autosave.grid(row=0, column=3,pady = 10, padx=10)

        # Confidence threshold
        conf_frame = tk.Frame(self.master, bg = "#326273")
        conf_frame.grid(row = 3, column = 0, sticky = "nsew", pady=5)
        # Create style object
        style = ttk.Style()

        # Configure the default Label font
        style.configure('TLabel', font=('Arial', 12))

        tk.Label(conf_frame, text="Confidence Threshold:", bg="#326273", foreground="white",
                                     font = ("Arial", 11)).grid(row = 0, column = 0)
        self.conf_var = tk.DoubleVar(value=0.5)
        # Configure the style for the scale
        style.configure("Custom.Horizontal.TScale",
                        background="#326273",  # Background color
                        foreground = "white",
                        troughcolor="#326273",  # Trough (track) color
                        bordercolor="#326273",  # Border color
                        darkcolor="#326273",  # Dark part of the groove
                        lightcolor="#326273")  # Light part of the groove
        conf_slider = ttk.Scale(conf_frame, from_=0.1, to=1, length = 760, variable=self.conf_var,
                               style= "Custom.Horizontal.TScale",
                                command=lambda e: self.update_conf_label())
        conf_slider.grid(row = 0, column = 1, sticky = "news", padx=5)

        self.conf_label = tk.Label(conf_frame, text="0.5", font = ("Arial", 11) )
        self.conf_label.grid(row = 0, column = 3, sticky = "news")

        #Frame for Right hand Panel
        self.right_hand_panel = tk.LabelFrame(self.master,
                                              bg="#326273",
                                              text="Defect Count Panel",
                                              font=('Arial', 14))
        self.right_hand_panel.grid(row = 0, column = 1,sticky = "nsew", pady=20, padx=20)
        # Load image with Pillow
        uzlogo = Image.open("uz logo.png")
        nrzlogo = Image.open("nrz logo.jpg")

        # Resize if needed
        img = uzlogo.resize((200, 100), Image.Resampling.LANCZOS)
        img2 = nrzlogo.resize((200, 100), Image.Resampling.LANCZOS)

        self.logo_image = ImageTk.PhotoImage(img)
        self.logo_image2 = ImageTk.PhotoImage(img2)

        self.canvas1 = tk.Canvas(self.right_hand_panel, width = 200, height = 100, bg="black")
        self.canvas1.grid(row=0, column=0, sticky="nws", padx=10, pady=10)
        self.canvas1.create_image(0, 0,
                                 image=self.logo_image,
                                 anchor="nw")
        self.canvas2 = tk.Canvas(self.right_hand_panel, width=200, height=100, bg="black")
        self.canvas2.grid(row=0, column=1, sticky="nws", padx=10, pady=10)
        self.canvas2.create_image(0, 0,
                                  image=self.logo_image2,
                                  anchor="nw")
        # Basic Defect number display
        self.defective_rail_count = 0
        self.defective_rail_label = ttk.Label(self.right_hand_panel, background="#326273",
                                         foreground="white", text =f"Defective Rails: {self.defective_rail_count}", font=('Arial', 11))
        self.defective_rail_label.grid(row = 1, column = 0, padx = 20, pady=2)

        self.poor_sleepers_count = 0
        self.poor_sleepers_label = ttk.Label(self.right_hand_panel, background="#326273", foreground="white",
                                        text=f"Poor Sleepers: {self.poor_sleepers_count}",font=('Arial', 11))
        self.poor_sleepers_label.grid(row=2, column=0, padx=20, pady=2)

        self.fair_sleepers_count = 0
        self.fair_sleepers_label = ttk.Label(self.right_hand_panel, background="#326273", foreground="white",
                                        text=f"Fair Sleepers: {self.fair_sleepers_count}",font=('Arial', 11))
        self.fair_sleepers_label.grid(row=3, column=0, padx=20, pady=2)

        self.good_sleepers_count = 0
        self.good_sleepers_label = ttk.Label(self.right_hand_panel, background="#326273",foreground="white",
                                        text=f"Good Sleepers: {self.good_sleepers_count}",font=('Arial', 11))
        self.good_sleepers_label.grid(row=4, column=0, padx=20, pady=2)

        self.missing_dogscrew_count = 0
        self.missing_dogscrew_label = ttk.Label(self.right_hand_panel, background="#326273", foreground="white",
                                           text=f"Missing Dog-screws: {self.missing_dogscrew_count}",font=('Arial', 11))
        self.missing_dogscrew_label.grid(row=1, column=1, padx=20, pady=2)

        self.defective_fasteners_count = 0
        self.defective_fasteners_label = ttk.Label(self.right_hand_panel, background="#326273", foreground="white",
                                              text = f"Defective Fasteners: {self.defective_fasteners_count}",font=('Arial', 11))
        self.defective_fasteners_label.grid(row=2, column=1, padx=20, pady=2)

        self.defective_joints_count = 0
        self.defective_joints_label = ttk.Label(self.right_hand_panel,background="#326273",foreground="white",
                                           text=f"Defective Joints: {self.defective_joints_count}",font=('Arial', 11))
        self.defective_joints_label.grid(row=3, column=1, padx=20, pady=2)

        repairing_status = self.is_repairing
        if repairing_status:
            status = "Repairing"
        else:
            status = "Inspecting"
        self.repairing_status_display = tk.Label(self.right_hand_panel,
                                           text=f"Repairing Status: {status:}",
                                            background="#326273", foreground="white",font=('Arial', 12))
        self.repairing_status_display.grid(row=4, column=1, padx=20, pady=5)

        cant = self.cant
        self.cant_label = tk.Label(self.right_hand_panel,
                               background="#326273", foreground="white", text=f"Track Cant: ___",font=('Arial', 11))
        self.cant_label.grid(row=5, column=0, padx=20, pady=5)

        gps = self.gps_coordinates
        self.gps_label = tk.Label(self.right_hand_panel, background="#326273", foreground="white",
                               text=f"GPS Coordinates: ___",font=('Arial', 11))
        self.gps_label.grid(row=5, column=1, padx=20, pady=5)

        gauge = self.gauge
        self.gauge_label = tk.Label(self.right_hand_panel, background="#326273", foreground="white",
                               text=f"Track Gauge: ___",font=('Arial', 11))
        self.gauge_label.grid(row=6, column=0, padx=20, pady=5)

        action = self.action
        self.action_label = tk.Label(self.right_hand_panel, background="#326273", foreground="white",
                                text=f"Action: ___", font=('Arial', 11))
        self.action_label.grid(row=6, column=1, padx=20, pady=5)

        self.demo_label = tk.Label(self.right_hand_panel, background="#326273", foreground="white",
                                     text=f"Demo: ___", font=('Arial', 11))
        self.demo_label.grid(row=7, column=0, padx=20, pady=5)

        self.distance_label = tk.Label(self.right_hand_panel, background="#326273", foreground="white",
                                     text=f"Distance: ___", font=('Arial', 11))
        self.distance_label.grid(row=7, column=1, padx=20, pady=5)

        btn_point_cloud = tk.Button(self.right_hand_panel, text="View Point Cloud", command=self.view_point_cloud,
                                 bg = "#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_point_cloud.grid(row=8, column=0, padx=20, pady=5)

        self.btn_second_window = tk.Button(self.right_hand_panel, text="Open Ultrasonic Scanner", command=self.ultrasonic_window,
                                     bg = "#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        self.btn_second_window.grid(row=8,column=1, padx=20, pady=5)
        btn_connect_cart = tk.Button(self.right_hand_panel, text="Connect Cart", command=self.connect_arduino,
                                      bg = "#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_connect_cart.grid(row=9, column=0, padx=20, pady=3)

        btn_move_cart = tk.Button(self.right_hand_panel, text="Move Cart",command=self.move_cart,
                                       bg = "#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_move_cart.grid(row=9,column=1, padx=20, pady=5)

        btn_stop_cart = tk.Button(self.right_hand_panel, text="Stop Cart", command=self.stop_cart,
                                   bg = "#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_stop_cart.grid(row=10, column=0, padx=20, pady=5)

        btn_reverse_cart = tk.Button(self.right_hand_panel, text="Reverse Cart", command=self.reverse_cart,
                                   bg = "#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_reverse_cart.grid(row=10, column=1, padx=20, pady=5)
        btn_calibrate = tk.Button(self.right_hand_panel, text="Calibrate Cart", command=self.calibrate_cart,
                                 bg = "#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_history = tk.Button(self.right_hand_panel, text="Inspection History", command=self.view_history,
                                bg="#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_history.grid(row=11, column=0, padx=20, pady=5)

        btn_calibrate.grid(row=11, column=1, padx=20, pady=5)
        btn_demonstrate = tk.Button(self.right_hand_panel, text="Run Demonstration", command=lambda: self.send_data('RunDemo'),
                                  bg="#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_demonstrate.grid(row=12, column=0, padx=20, pady=5)
        btn_report = tk.Button(self.right_hand_panel, text="Print Report", command=self.report,
                                    bg="#326273", foreground="white", width=20, height=1, font=('Arial', 11))
        btn_report.grid(row=12, column=1, padx=20, pady=5)

        # ttk.Label(self.right_hand_panel, text="Move Mode:").grid(row=7, column=0, sticky="w", pady=2)
        # self.move_mode_var = tk.StringVar()
        # ttk.Combobox(self.right_hand_panel, textvariable=self.move_mode_var,
        #              values=["MoveScan", "MoveRepair", "MoveCruse", "MoveStop", "MoveReverse"]).grid(row=7, column=1, sticky="ew",
        #                                                                                    pady=2)


        # Frame for Bottom Right hand Panel
        self.operation_mode_panel = tk.LabelFrame(self.master, bg="#326273",
                                                  text="Operation Mode Selection",
                                              font=('Arial', 14))
        self.operation_mode_panel.grid(row=2, column=1, sticky="nsew", pady=20, padx=20)

        # self.mode1_var = tk.BooleanVar(value=True) # Inspect only
        # self.mode2_var = tk.BooleanVar(value=False) #Inspect and Repair

        # chk_mode1 = ttk.Checkbutton(self.operation_mode_panel, text="Inspect Only", variable=self.mode1_var, style = "TCheckbutton")
        # chk_mode1.grid(row=0, column =0, sticky="w") #,pady = 10, padx=10)
        #
        # chk_mode2 = ttk.Checkbutton(self.operation_mode_panel, text="Inspect and Repair", variable=self.mode2_var, style = "TCheckbutton")
        # chk_mode2.grid(row = 0, column = 1, pady = 10, padx=10)#, sticky=tk.W)

        ttk.Label(self.operation_mode_panel,background="#326273", text="Operation Mode:").grid(row=0, column=0, sticky="w", pady=5, padx=5)
        self.operation_mode_var = tk.StringVar()
        combo = ttk.Combobox(self.operation_mode_panel, background="#326273", textvariable=self.operation_mode_var,
                     values=["None", "Cruse","Inspect Only", "Inspect and Repair"])
        combo.set("None")
        combo.grid(row=0, column=1,sticky="ew", pady=5, padx=5)

        ttk.Label(self.operation_mode_panel,background="#326273", text="Route:").grid(row=0, column=2, sticky="w", pady=2, padx=5)
        self.origin_var = tk.StringVar()
        route_combo = ttk.Combobox(self.operation_mode_panel, background="#326273", textvariable=self.origin_var,
                             values=["None", "Shamva to Harare", "Harare to Mvuma", "Mvuma to Ngundu", "Ngundu to Beitbridge"])
        route_combo.set("None")
        route_combo.grid(row=0, column=3, sticky="ew", pady=2, padx=5)

        #Control Buttons
        last_frame = tk.Frame(self.master, bg = "#326273")
        last_frame.grid(row=3, column=1, sticky="news", padx=10, pady=10)

        btn_close_page = tk.Button(last_frame, text="Exit", command=self.on_closing, bg = "#326273", foreground="white",
                                   width=15, height=1, font=('Arial', 11))
        btn_close_page.grid(row=0, column=3, pady=10, padx=8)

        btn_reset_page = tk.Button(last_frame, text="Reset", command=self.reset_page,bg = "#326273", foreground="white",
                                   width=15, height=1, font=('Arial', 11))
        btn_reset_page.grid(row=0, column=2, pady=10, padx=8)

        btn_back_page = tk.Button(last_frame, text="Back", command=self.back_page, bg = "#326273", foreground="white",
                                  width=15, height=1, font=('Arial', 11))
        btn_back_page.grid(row=0, column=0, pady=10, padx=8)

        # Dictionary to easily map received labels to their corresponding Tkinter widgets
        self.display_labels = {
            "GPS": self.gps_label,
            "Cant": self.cant_label,
            "Gauge": self.gauge_label,
            "Action": self.action_label,
            "Demo": self.demo_label,
            "Distance": self.distance_label
        }

        # Schedule the periodic check for new data from the serial thread
        self.master.after(100, self.process_serial_data)


    def reset_page(self):

        # Ensure proper shutdown of the serial thread and port
        self.running = False
        self.stop_video()
        if self.cap is not None:
            self.cap.release()
        if self.serial_thread:
            self.serial_thread.join(timeout=1)  # Give the thread a moment to close
        if self.ser and self.ser.is_open:
            self.ser.close()
        #self.master.destroy()
        for widget in self.master.winfo_children():
            widget.destroy()
        self.create_widgets()
    def back_page(self):
        return None

    def view_history(self):
        # Switch from Detection output to History label
        for widget in self.video_frame.winfo_children():
            widget.destroy()
        self.video_frame.destroy()
        for widget in self.control_frame.winfo_children():
            widget.destroy()


        # Frame for History display
        self.gallery_frame = tk.LabelFrame(self.master, font=('Arial', 14), text="History Output")
        self.gallery_frame.grid(row=0, column=0, padx=10, pady=10)

        self.close_history = ttk.Button(self.control_frame, text="Close History", command=self.close_history, style = "LargeFont.TButton")
        self.close_history.grid(row=1, column=1)

        self.scrollbar = tk.Scrollbar(self.gallery_frame)
        self.scrollbar.grid(row=0, column=1, sticky="news")

        self.canvas = tk.Canvas(self.gallery_frame, bg='black', width=840, height=500)
        self.canvas.grid(row=0, column=0, sticky="nws", padx=10, pady=10)

        self.inner_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.inner_frame, anchor=tk.NW)

        self.inner_frame.bind("<Configure>", lambda e: self.canvas.configure(
            scrollregion=self.canvas.bbox("all")))

        # Load thumbnails from database
        images = self.db.get_all_images()

        for img_id, timestamp, desc, category, thumbnail in images:
            frame = ttk.Frame(self.inner_frame)
            frame.grid(row=0, column = 0, pady=5)

            # Display thumbnail
            img = Image.open(io.BytesIO(thumbnail))
            photo = ImageTk.PhotoImage(img)

            label = ttk.Label(frame, image=photo)
            label.image = photo  # Keep reference
            label.grid(row = 0, column = 0)

            # Display metadata
            meta_frame = ttk.Frame(frame)
            meta_frame.grid(row=0, column=0, padx=10, pady=10)

            ttk.Label(meta_frame, text=f"Time: {timestamp[:19]}").grid(row=0, column=1)
            ttk.Label(meta_frame, text=f"Desc: {desc}").grid(row =1, column =1)
            ttk.Label(meta_frame, text=f"Category: {category}").grid(row=2, column =1)

            # Add view button
            ttk.Button(frame, text="View",
                       command=lambda i=img_id: self.view_image(i)).grid(row =0, column=2)
    def view_image(self, image_id):
        image_data = self.db.get_full_image(image_id)
        if image_data:
            top = tk.Toplevel(self.master)
            top.title("Full Image View")

            img = Image.open(io.BytesIO(image_data))
            photo = ImageTk.PhotoImage(img)

            label = ttk.Label(top, image=photo)
            label.image = photo  # Keep reference
            label.grid()
        else:
            messagebox.showerror("Error", "Image not found")

    def close_history(self):
        for widget in self.master.winfo_children():
            widget.destroy()
        self.create_widgets()

    def view_point_cloud(self):
        if self.is_realsense_d435():
            print("RealSense D435 camera detected!")
            point_cloud = self.get_point_cloud()
            print(f"Point cloud shape: {point_cloud.shape}")
            print("First 5 points:")
            print(point_cloud[:5])
        else:
            messagebox.showerror("Error", "No RealSense D435 camera found")

    def is_realsense_d435(self):
        try:
            ctx = rs.context()
            if len(ctx.devices) == 0:
                return False

            for dev in ctx.devices:
                if 'D435' in dev.get_info(rs.camera_info.name):
                    return True
            return False
        except:
            return False

    def get_point_cloud(self):
        # Configure depth and color streams
        pipeline = rs.pipeline()
        config = rs.config()

        # Enable both depth and color streams
        config.enable_stream(rs.stream.depth, 640, 480, rs.format.z16, 30)
        config.enable_stream(rs.stream.color, 640, 480, rs.format.bgr8, 30)

        # Start streaming
        profile = pipeline.start(config)

        # Get depth sensor's depth scale
        depth_sensor = profile.get_device().first_depth_sensor()
        depth_scale = depth_sensor.get_depth_scale()

        # Create align object to align depth to color
        align_to = rs.stream.color
        align = rs.align(align_to)

        try:
            while True:
                # Wait for a coherent pair of frames: depth and color
                frames = pipeline.wait_for_frames()

                # Align the depth frame to color frame
                aligned_frames = align.process(frames)

                # Get aligned frames
                depth_frame = aligned_frames.get_depth_frame()
                color_frame = aligned_frames.get_color_frame()

                if not depth_frame or not color_frame:
                    continue

                # Convert images to numpy arrays
                depth_image = np.asanyarray(depth_frame.get_data())
                color_image = np.asanyarray(color_frame.get_data())

                # Create point cloud
                pc = rs.pointcloud()
                points = pc.calculate(depth_frame)

                # Get the vertices of the point cloud
                vtx = np.asanyarray(points.get_vertices())
                vertices = np.stack([vtx['f0'], vtx['f1'], vtx['f2']], axis=-1)

                # Apply colormap on depth image
                depth_colormap = cv2.applyColorMap(
                    cv2.convertScaleAbs(depth_image, alpha=0.03),
                    cv2.COLORMAP_JET
                )

                # Stack images horizontally
                images = np.hstack((color_image, depth_colormap))

                # Show images
                cv2.namedWindow('RealSense', cv2.WINDOW_AUTOSIZE)
                cv2.imshow('RealSense', images)

                # Press 'q' to quit
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break

                return vertices

        finally:
            pipeline.stop()
            cv2.destroyAllWindows()

        #return vertices

    def livestream(self):
        cam_port = 0
        user_input = simpledialog.askinteger("Input", "Enter Camera Port:", parent=root, initialvalue=0)
        cam_port = user_input
        self.cap = cv2.VideoCapture(cam_port)
        self.is_running = True
        #self.play_stream(1)
        self.video_loop()

    def video_loop(self):
        if not self.is_running:
            return

        # Check if camera opened successfully
        if not self.cap.isOpened():
            print("Error: Could not open video source")
            return

        try:
            # Read frame
            ret, frame = self.cap.read()
            if not ret:
                print("Error: Failed to capture frame")
                self.stop_video()
                return

            # Store the current frame for possible saving
            self.current_frame = frame.copy()

            # Perform inference
            results = self.seg_model1(frame, conf=self.conf_var.get(), verbose=False)

            # Get annotated frame
            if len(results) > 0:
                result = results[0]
                annotated_frame = result.plot()

                # Display class information
                if result.boxes is not None:
                    # SnapShoot the defect
                    self.autosave()
                    for box, cls in zip(result.boxes, result.boxes.cls):
                        class_id = int(cls)
                        class_name = self.seg_model1.names[class_id]
                        confidence = float(box.conf[0])
                        if confidence >= self.conf_var.get():
                            #print(f"Detected: {class_name} ({confidence:.2f})")
                            # Add count and action
                            if class_name == "Defective Joint":
                                self.class_variable = class_name
                                self.defective_joints_count += 1
                                self.defective_joints_label.config(
                                    text=f"Defective Joints: {self.defective_joints_count}")
                            elif class_name == "Missing Dogscrew":
                                self.class_variable = class_name
                                self.missing_dogscrew_count += 1
                                self.missing_dogscrew_label.config(
                                    text=f"Missing Dogscrew: {self.missing_dogscrew_count}")
                                # first check compatibility with depth camera to get coordinates
                                if self.operation_mode_var == "Inspect and Repair":
                                    self.is_repairing = True
                                    xywh = results.boxes.xywh  # x_center, y_center, width, height
                                    x_center_pixels, y_center_pixels, width_pixels, height_pixels = xywh.flatten()
                                    # Camera Parameteres
                                    z = [479, 441, 368, 332, 209, 277, 250, 223, 209, 199]  # height_pixels
                                    x = [224, 186, 169, 155, 136, 131, 125, 102, 100, 89]  # pixel width
                                    y = [25, 30, 35, 40, 45, 50, 55, 60, 65, 70]  # Real Distance in CM

                                    coff = np.polyfit(x, y, 2)  # y = Ax**2 + Bx + C
                                    A, B, C = coff
                                    depth = width_pixels
                                    depthCM = (A * (depth ** 2)) + (B * depth) + C  # depth to the crewing point
                                    depthCM = round(depthCM, 2)
                                    #Define camera parameters
                                    focal_length = 2000  # in mm
                                    sensor_size = (23.5, 15.6)  # in mm
                                    resolution = (640, 480)  # in pixels
                                    camera_params = {
                                            'focal_length': 2000,  # in pixels
                                            'sensor_size': (23.5, 15.6),  # in mm
                                            'resolution': (640, 480)  # in pixels
                                        }
                                    # Calculate x, y coordinates in pixels
                                    xrel = x_center_pixels - resolution[0] / 2
                                    yrel = y_center_pixels - resolution[1] / 2
                                    # Convert to real World Coordinates in mm
                                    x_real = ((xrel * depthCM) / focal_length) * 10
                                    y_real = ((yrel * depthCM) / focal_length) * 10
                                    depth_mm = depthCM * 10
                                    #self.send_repair_command(x_real, y_real, depth_mm)
                            elif class_name == "Missing Fastener":
                                self.class_variable = class_name
                                self.defective_fasteners_count += 1
                                self.defective_fasteners_label.config(
                                    text=f"Defective Fasteners: {self.defective_fasteners_count}")
                            # elif class_name == "Good Fasteners":
                            #     print("")
                            # Save in database
                            if hasattr(self, 'current_frame') and self.current_frame is not None:
                                    img = Image.fromarray(self.current_frame)
                                    img_bytes = io.BytesIO()
                                    img.save(img_bytes, format='JPEG')
                                    img_data = img_bytes.getvalue()
                                    #self.db.save_image(img_data, self.class_variable, self.gps_variable,
                                                       #self.distance_variable, self.action_variable)


            else:
                annotated_frame = frame

            # Convert to RGB and resize for display
            annotated_frame = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(annotated_frame)

            # Get canvas dimensions
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()

            # Maintain aspect ratio
            img_ratio = img.width / img.height
            canvas_ratio = canvas_width / canvas_height

            if canvas_ratio > img_ratio:
                new_height = canvas_height
                new_width = int(new_height * img_ratio)
            else:
                new_width = canvas_width
                new_height = int(new_width / img_ratio)

            img = img.resize((new_width, new_height), Image.LANCZOS)
            self.tk_img = ImageTk.PhotoImage(image=img)

            # Update canvas
            self.canvas.delete("all")
            self.canvas.create_image(
                canvas_width // 2, canvas_height // 2,
                anchor=tk.CENTER,
                image=self.tk_img
            )

            # Schedule next frame update if still running
            if self.is_running:
                self.master.after(10, self.video_loop)

        except Exception as e:
            print(f"Error in video loop: {e}")
            self.stop_video()
    def send_repair_command(self, x_center, y_center, height,):
        """Send data over serial if connected."""
        if not self.ser or not self.ser.is_open:
            messagebox.showerror("Error", "Not connected to serial port!")
            return
        try:
            # Create a string with the command and values separated by commas
            data_string = f"repair,{x_center},{y_center}, {height}\n"

            # Send the data to Arduino
            self.ser.write(data_string.encode())
            print(f"Sent: {data_string.strip()}")

            # raise flag for repairing
            self.is_repairing = True
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Failed to send data: {e}")

    def move_cart(self):
        data_string = ""
        if self.operation_mode_var.get() == "Inspect Only":
            data_string = "MoveScan"
            self.move_scan = True
        if self.operation_mode_var.get() == "Inspect and Repair":
            data_string = "MoveRepair"
            self.move_repair = True
        if self.operation_mode_var.get() == "Cruse":
            data_string = "MoveCruse"
            self.move_cruse = True

        # send the data to Arduino
        if not self.ser or not self.ser.is_open:
            messagebox.showerror("Error", "Not connected to serial port!")
            return
        try:
            self.ser.write(data_string.encode())
            print(f"Sent: {data_string.strip()}")
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Failed to send data: {e}")
    def update_gps_cant_gauge(self):
        label, coordinates = self.read_arduino_data()
        if label == "GPS":
            self.gps_coordinates = coordinates
        elif label == "Gauge":
            self.gauge = coordinates
        elif label == "Cant":
            self.cant = coordinates

    def stop_cart(self):
        self.move_stop = True
        data_string = "MoveStop"
        # send the data to Arduino
        if not self.ser or not self.ser.is_open:
            messagebox.showerror("Error", "Not connected to serial port!")
            return
        try:
            self.ser.write(data_string.encode())
            print(f"Sent: {data_string.strip()}")
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Failed to send data: {e}")

    def reverse_cart(self):
        self.is_move_reverse = True
        data_string = "MoveReverse"
        # send the data to Arduino
        if not self.ser or not self.ser.is_open:
            messagebox.showerror("Error", "Not connected to serial port!")
            return
        try:
            self.ser.write(data_string.encode())
            print(f"Sent: {data_string.strip()}")
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Failed to send data: {e}")

    def demonstrate(self):
        data_string = "send"
        # send the data to Arduino
        if not self.ser or not self.ser.is_open:
            messagebox.showerror("Error", "Not connected to serial port!")
            return
        try:
            self.ser.write(data_string.encode('utf-8'))
            print(f"Sent: {data_string.strip()}")
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Failed to send data: {e}")

    def send_data(self, data):
        if self.ser and self.ser.is_open:
            try:
                self.ser.write(data.encode('utf-8'))
                print(f"Sent: {data}")
            except serial.SerialException as e:
                print(f"Error sending data: {e}")
                # self.data_label.config(text=f"Send Error: {e}")
        else:
            messagebox.showerror("Error", f"Failed to send data: ")
            return
            # self.data_label.config(text="Not connected to Arduino.")

    def report(self):
        # Connect to the database
        conn = sqlite3.connect(DATABASE_FILE)

        # Create output directory for images
        os.makedirs(IMAGE_OUTPUT_DIR, exist_ok=True)

        # Read data from database
        query = "SELECT timestamp, defect_class, gps_coordinates, distance, gauge, cant, action_taken, image_data FROM images"
        df = pd.read_sql_query(query, conn)

        # Close connection
        conn.close()

        # Convert timestamp to datetime
        df['timestamp'] = pd.to_datetime(df['timestamp'])

        # Generate defect count by class
        defect_counts = df['defect_class'].value_counts().reset_index()
        defect_counts.columns = ['Defect Class', 'Count']
        origin = self.origin_var.get()

        # Create HTML report
        html_content = f"""
            <html>
            <head>
                <title>Defect Inspection Report</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ color: #2c3e50; }}
                    h2 {{ color: #34495e; margin-top: 30px; }}
                    table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                    th {{ background-color: #f2f2f2; }}
                    tr:nth-child(even) {{ background-color: #f9f9f9; }}
                    .summary {{ background-color: #e8f4f8; padding: 15px; border-radius: 5px; }}
                    .image-container {{ margin: 10px 0; }}
                    .image-caption {{ font-size: 0.9em; color: #666; }}
                </style>
            </head>
            <body>
                <h1>Defect Inspection Report</h1>
                <div class="summary">
                    <p><strong>Report generated on:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    <p><strong>Total defects recorded:</strong> {len(df)}</p>
                </div>

                <h2>Defect Count by Class</h2>
                {defect_counts.to_html(index=False)}

                <h2>All Defect Records</h2>
                <table>
                    <tr>
                        <th>Timestamp</th>
                        <th>Defect Class</th>
                        <th>GPS Coordinates</th>
                        <th>Track Gauge</th>
                        <th>Track Cant</th>
                        <th>Action Taken</th>
                        <th>Distance from {origin}</th>
                        <th>Image Preview</th>
                    </tr>
            """

        # Add each defect record to the report
        for idx, row in df.iterrows():
            # Save image to file and create HTML reference
            img_filename = f"defect_{idx}.jpg"
            img_path = os.path.join(IMAGE_OUTPUT_DIR, img_filename)

            try:
                if row['image_data']:
                    image = Image.open(io.BytesIO(row['image_data']))
                    # Resize for thumbnail
                    image.thumbnail((200, 200))
                    image.save(img_path)

                    img_tag = f'<div class="image-container"><img src="{img_path}" width="200"><br><span class="image-caption">Image {idx}</span></div>'
                else:
                    img_tag = "No image available"
            except Exception as e:
                img_tag = f"Image unavailable ({str(e)})"

            html_content += f"""
                    <tr>
                        <td>{row['timestamp']}</td>
                        <td>{row['defect_class']}</td>
                        <td>{row['gps_coordinates']}</td>
                        <td>{row['gauge']}</td>
                        <td>{row['cant']}</td>
                        <td>{row['action_taken']}</td>
                        <td>{row['distance']}</td>
                        <td>{img_tag}</td>
                    </tr>
                """

        html_content += """
                </table>
            </body>
            </html>
            """

        # Save the report
        with open(REPORT_FILE, 'w') as f:
            f.write(html_content)
        messagebox.showinfo("Defect Report ", f"Report generated successfully: {REPORT_FILE}")
        print(f"Report generated successfully: {REPORT_FILE}")

        # Optional: Generate defect location plot if GPS coordinates are valid
        try:
            # Extract latitude and longitude from GPS coordinates
            # Assuming format "lat,lon" or similar
            df[['latitude', 'longitude']] = df['gps_coordinates'].str.split(',', expand=True).astype(float)

            plt.figure(figsize=(10, 8))
            for defect_class in df['defect_class'].unique():
                subset = df[df['defect_class'] == defect_class]
                plt.scatter(subset['longitude'], subset['latitude'], label=defect_class, alpha=0.7)

            plt.title('Defect Locations')
            plt.xlabel('Longitude')
            plt.ylabel('Latitude')
            plt.legend()
            plt.grid()
            plt.tight_layout()
            plt.savefig('defect_locations.png')
            messagebox.showinfo("Visualization Picture", f"Defect location map generated: defect_locations.png")
            print("Defect location map generated: defect_locations.png")
        except Exception as e:
            messagebox.showinfo("Visualization Picture", f"Could not generate location map")
            print(f"Could not generate location map: {str(e)}")

    def calibrate_cart(self):
        data_string = "Calibrate"
        # send the data to Arduino
        if not self.ser or not self.ser.is_open:
            messagebox.showerror("Error", "Not connected to serial port!")
            return
        try:
            self.ser.write(data_string.encode())
            print(f"Sent: {data_string.strip()}")
        except serial.SerialException as e:
            messagebox.showerror("Error", f"Failed to send data: {e}")
    def find_arduino_port(self):
        """Identify Arduino port automatically"""
        ports = serial.tools.list_ports.comports()
        for port in ports:
            if 'Arduino' in port.description or 'USB Serial Device' in port.description:
                return port.device
        return None
    def connect_arduino(self):
        try:
            # Open serial port with a timeout
            self.ser = serial.Serial(self.serial_port, self.baud_rate, timeout=1)
            time.sleep(2)  # Give Arduino time to reset and initialize
            print(f"Connected to Arduino on {self.serial_port}")
            messagebox.showinfo("Serial Connection", f"Connected to {self.serial_port}!")
            self.repairing_status_display.config(text="Status: Connected", fg="green")
            self.running = True

            # Start a separate thread for reading serial data
            self.serial_thread = threading.Thread(target=self.read_serial_data, daemon=True)
            self.serial_thread.start()

            # self.connect_button.config(state=tk.DISABLED)
            # self.disconnect_button.config(state=tk.NORMAL)
            # self.send_button.config(state=tk.NORMAL)

        except serial.SerialException as e:
            print(f"Could not open serial port: {e}")
            # self.status_label.config(text=f"Error: {e}", fg="red")

    def disconnect_arduino(self):
        if self.ser and self.ser.is_open:
            self.running = False  # Signal the thread to stop
            if self.serial_thread:
                self.serial_thread.join(timeout=1)  # Wait for thread to finish (with a timeout)
            self.ser.close()
            print("Disconnected from Arduino")
        self.repairing_status_display.config(text="Status: Disconnected", fg="red")
        # self.connect_button.config(state=tk.NORMAL)
        # self.disconnect_button.config(state=tk.DISABLED)
        # Reset displayed values
        for label_text, widget in self.display_labels.items():
            widget.config(text=f"{label_text}: --")
    def read_serial_data(self):
        # This function runs in a separate thread
        while self.running:
            try:
                if self.ser and self.ser.is_open and self.ser.in_waiting > 0:
                    line = self.ser.readline().decode('utf-8').strip()
                    if line:  # Only process non-empty lines
                        self.data_queue.put(line)  # Put data into the queue
                # Small delay to prevent busy-waiting and release CPU
                time.sleep(0.01)
            except serial.SerialException as e:
                print(f"Serial read error: {e}")
                self.running = False  # Stop the thread on error
                self.data_queue.put(f"ERROR: {e}")  # Signal error to GUI
            except Exception as e:
                print(f"An unexpected error occurred in serial read thread: {e}")
                self.running = False
                self.data_queue.put(f"THREAD_ERROR: {e}")
    def process_serial_data(self):
        # This function runs in the main Tkinter thread
        while not self.data_queue.empty():
            raw_data = self.data_queue.get()

            # Handle error messages coming from the serial thread
            if raw_data.startswith("ERROR:") or raw_data.startswith("THREAD_ERROR:"):
                #self.status_label.config(text=raw_data, fg="red")
                self.disconnect_arduino()  # Automatically disconnect on critical errors
                break  # Stop processing further queued items for now

            # Attempt to parse the "Label: Value" format
            if ":" in raw_data:
                try:
                    label, value = raw_data.split(":", 1)  # Split only on the first colon
                    label = label.strip()
                    value = value.strip()

                    # Update the specific Tkinter label for the received data type
                    if label in self.display_labels:
                        # Store the value in the corresponding variable
                        if label == "Gauge":
                            self.gauge_variable = value
                        elif label == "Distance":
                            self.distance_variable = value
                        elif label == "Cant":
                            self.cant_variable = value
                        elif label == "GPS":
                            self.gps_variable = value
                        elif label == "Action":
                            self.action_variable = value
                        elif label == "Demo":
                            self.demo_variable = value
                            print(self.demo_variable)
                            if self.demo_variable == "Skewed Sleeper":
                                self.class_variable = "Skewed Sleeper"
                            elif self.demo_variable == "Missing Dogscrew":
                                self.class_variable = "Missing Dogscrew"
                            elif self.demo_variable == "Cracked Sleeper":
                                self.class_variable = "Cracked Sleeper"
                            elif self.demo_variable == "Missing Fastener":
                                self.class_variable = "Missing Fastener"
                            elif self.demo_variable == "Defective Rail":
                                self.class_variable = "Defective Rail"

                        print("Gauge: ", self.gauge_variable)
                        print("Distance: ", self.distance_variable)
                        print("Cant: ", self.cant_variable)
                        print("GPS: ", self.gps_variable)
                        print("Action: ", self.action_variable)
                        print("Demo: ", self.demo_variable)


                        # Update the label display
                        self.display_labels[label].config(text=f"{label}: {value}")

                        # Store the current frame for possible saving
                        #self.cap = cv2.VideoCapture(cam_port)
                        self.is_running = True
                        ret, frame = self.cap.read()
                        self.current_frame = frame.copy()

                        # Save in database
                        if hasattr(self, 'current_frame') and self.current_frame is not None:
                            img = Image.fromarray(self.current_frame)
                            img_bytes = io.BytesIO()
                            img.save(img_bytes, format='JPEG')
                            img_data = img_bytes.getvalue()
                            self.db.save_image(img_data, self.class_variable, self.gps_variable, self.distance_variable,
                                               self.gauge_variable, self.cant_variable, self.action_variable)

                        #Save in database
                        # Convert to JPEG bytes
                        # img = Image.fromarray(self.current_frame)
                        # img_bytes = io.BytesIO()
                        # img.save(img_bytes, format='JPEG')
                        # img_data = img_bytes.getvalue()
                        # self.db.save_image(img_data, self.class_variable, self.gps_variable, self.distance_variable, self.action_variable)

                    else:
                        print(f"Warning: Received unknown label '{label}'. Value: '{value}'")
                        # You could update a general status label here if desired
                        self.repairing_status_display.config(text=f"Unknown Data: {label}:{value}", fg="orange")

                except ValueError:
                    # This happens if a colon is found but split doesn't result in two parts
                    print(f"Parsing error: '{raw_data}' - unexpected format.")
                    self.repairing_status_display.config(text=f"Parsing Error: {raw_data}", fg="orange")
            else:
                # Handle data without a colon (e.g., "Arduino Ready" message)
                print(f"Received non-labeled data: '{raw_data}'")
                if "Arduino Ready" in raw_data:  # Example: specifically handle a "ready" message
                    self.repairing_status_display.config(text="Status: Arduino Ready", fg="green")
                else:
                    self.repairing_status_display.config(text=f"Data: {raw_data}", fg="blue")  # Display general messages

        self.master.after(100, self.process_serial_data)  # Schedule itself again
    def on_closing(self):
        # Ensure proper shutdown of the serial thread and port
        self.running = False
        self.stop_video()
        if self.cap is not None:
            self.cap.release()
        if self.serial_thread:
            self.serial_thread.join(timeout=1)  # Give the thread a moment to close
        if self.ser and self.ser.is_open:
            self.ser.close()
        self.master.destroy()
    def parse_arduino_data(self, line):
        """Parse formatted data from Arduino"""
        try:
            # Split into label and values parts
            label, values_str = line.strip().split(':')
            # Split values into individual numbers
            values = [float(x) for x in values_str.split(',')]
            return label, values
        except Exception as e:
            print(f"Error parsing line: {line} - {str(e)}")
            return None, None
    def read_arduino_data(self):
        try:
            while True:
                if self.ser.in_waiting > 0:
                    line = self.ser.readline().decode('utf-8').strip()
                    print(line)
                    if line:
                        label, values = self.parse_arduino_data(line)
                        if label and values:
                            print(f"Received {label}: {values}")
                            return label, values
                time.sleep(1)

        except KeyboardInterrupt:
            print("\nClosing serial connection")
            self.ser.close()
    def update_conf_label(self):
        self.conf_label.config(text=f"{self.conf_var.get():.1f}")
    def autosave(self):
        if self.autosave_var.get():

            if self.current_frame is None:
                return

            # Convert to JPEG bytes
            img = Image.fromarray(self.current_frame)
            img_bytes = io.BytesIO()
            img.save(img_bytes, format='JPEG')
            img_data = img_bytes.getvalue()

            # Save to database
            self.db.save_image(img_data, self.class_variable, self.gps_variable, self.distance_variable,
                               self.gauge_variable, self.cant_variable, self.action_variable)
    def open_video(self):
        self.video_path = filedialog.askopenfilename(filetypes=[("Video Files", "*.mp4 *.avi *.mov")])
        if self.video_path:
            self.cap = cv2.VideoCapture(self.video_path)
            self.show_frame()

    def start_video(self):
        if self.cap is not None:
            self.is_running = True
            self.process_video()

    def stop_video(self):
        self.is_running = False

    def process_video(self):
        if self.is_running and self.cap is not None:
            ret, frame = self.cap.read()
            if ret:
                processed_frame = self.process_frame(frame)
                self.show_processed_frame(processed_frame)
                self.master.after(1, self.process_video)
            else:
                self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                self.master.after(1, self.process_video)

    def show_frame(self):
        if self.cap is not None:
            ret, frame = self.cap.read()
            if ret:
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                self.current_frame = frame
                self.update_canvas(frame)

    def take_snapshot(self):
        if self.current_frame is not None:
            file_path = filedialog.asksaveasfilename(defaultextension=".jpg",
                                                     filetypes=[("JPEG Image", "*.jpg")])
            if file_path:
                img = Image.fromarray(self.current_frame)
                img.save(file_path)
            img = Image.fromarray(self.current_frame)
            img_bytes = io.BytesIO()
            img.save(img_bytes, format='JPEG')
            img_data = img_bytes.getvalue()
            self.db.save_image(img_data, self.class_variable, self.gps_variable, self.distance_variable,
                               self.gauge_variable, self.cant_variable, self.action_variable)

    def open_image(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.jpg *.png *.jpeg")])
        if file_path:
            img = cv2.imread(file_path)
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

            #perform detections
            # Create blank mask for combined segmentation
            combined_mask = np.zeros(img.shape[:2], dtype=np.uint8)
            if self.seg1_var.get():
                seg_results1 = self.seg_model1(img, conf=self.conf_var.get(), verbose=False)[0]
                if seg_results1.masks is not None:
                    for mask, cls_id in zip(seg_results1.masks.data.cpu().numpy(),
                                            seg_results1.boxes.cls.cpu().numpy()):
                        mask_resized = cv2.resize(mask, (img.shape[1], img.shape[0]))
                        combined_mask = np.maximum(combined_mask, (mask_resized * 255).astype(np.uint8))

                        # Get class name and draw label
                        class_name = self.seg_model1.names[int(cls_id)]
                        y, x = np.unravel_index(np.argmax(mask_resized), mask_resized.shape)
                        cv2.putText(img, class_name, (x, y), cv2.FONT_HERSHEY_SIMPLEX,
                                    0.5, (255, 0, 0), 1)
                        if class_name is not None:
                            self.class_variable = class_name
                            # Add count and action
                            if class_name == "Defective Joint":
                                self.defective_joints_count += 1
                                self.defective_joints_label.config(
                                    text=f"Defective Joints: {self.defective_joints_count}")
                            elif class_name == "Missing Dogscrew":
                                self.missing_dogscrew_count += 1
                                self.missing_dogscrew_label.config(
                                    text=f"Missing Dogscrew: {self.missing_dogscrew_count}")
                            elif class_name == "Missing Fastener":
                                self.defective_fasteners_count += 1
                                self.defective_fasteners_label.config(
                                    text=f"Defective Fasteners: {self.defective_fasteners_count}")


            if self.seg2_var.get():
                seg_results2 = self.seg_model2(img, conf=self.conf_var.get(), verbose=False)[0]
                if seg_results2.masks is not None:
                    for mask, cls_id in zip(seg_results2.masks.data.cpu().numpy(),
                                            seg_results2.boxes.cls.cpu().numpy()):
                        mask_resized = cv2.resize(mask, (img.shape[1], img.shape[0]))
                        combined_mask = np.maximum(combined_mask, (mask_resized * 255).astype(np.uint8))

                        # Get class name and draw label
                        class_name = self.seg_model2.names[int(cls_id)]
                        y, x = np.unravel_index(np.argmax(mask_resized), mask_resized.shape)
                        cv2.putText(img, class_name, (x, y), cv2.FONT_HERSHEY_SIMPLEX,
                                    0.5, (255, 0, 0), 1)
                        if class_name is not None:
                            self.class_variable = class_name
                            # Add count
                            if class_name == "defective-rail":
                                self.defective_rail_count += 1
                                self.defective_rail_label.config(text=f"Defective Rails: {self.defective_rail_count}")


            # Apply color overlay for segmentation masks
            colored_mask = cv2.applyColorMap(combined_mask, cv2.COLORMAP_JET)
            overlay = cv2.addWeighted(img, 0.5, colored_mask, 0.3, 0)

            # Process OBB detections if selected
            if self.obb_var.get():
                obb_results = self.obb_model(img, conf=self.conf_var.get(), verbose=False)[0]
                if obb_results.obb is not None:
                    for box, cls_id in zip(obb_results.obb.xyxyxyxy.cpu().numpy(), obb_results.obb.cls.cpu().numpy()):
                        pts = box.reshape(-1, 2).astype(np.int32)
                        cv2.polylines(overlay, [pts], isClosed=True, color=(0, 255, 255), thickness=2)

                        # Get class name and draw label
                        class_name = self.obb_model.names[int(cls_id)]
                        cv2.putText(overlay, class_name, (pts[0][0], pts[0][1] - 10),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)
                        if class_name is not None:
                            self.class_variable = class_name
                            # Add count
                            if class_name == "Fair Conrete Sleeper":
                                self.fair_sleepers_count += 1
                                self.fair_sleepers_label.config(text=f"Fair Sleepers: {self.fair_sleepers_count}")
                            elif class_name == "Good Concrete Sleeper":
                                self.good_sleepers_count += 1
                                self.good_sleepers_label.config(text=f"Good Sleepers: {self.poor_sleepers_count}")
                            elif class_name == "Poor Concrete Sleeper":
                                self.poor_sleepers_count += 1
                                self.poor_sleepers_label.config(text=f"Poor Sleepers: {self.poor_sleepers_count}")

            self.current_frame = overlay
            self.update_canvas(overlay)

    def detect_image(self):
        if self.current_frame is not None:
            processed_frame = self.process_frame(self.current_frame)
            self.show_processed_frame(processed_frame)

    def process_frame(self, frame):
        # Create blank mask for combined segmentation
        combined_mask = np.zeros(frame.shape[:2], dtype=np.uint8)

        # Process with selected models
        if self.seg1_var.get():
            seg_results1 = self.seg_model1(frame, conf=self.conf_var.get(), verbose=False)[0]
            if seg_results1.masks is not None:
                for mask, cls_id in zip(seg_results1.masks.data.cpu().numpy(), seg_results1.boxes.cls.cpu().numpy()):
                    mask_resized = cv2.resize(mask, (frame.shape[1], frame.shape[0]))
                    combined_mask = np.maximum(combined_mask, (mask_resized * 255).astype(np.uint8))

                    # Get class name and draw label
                    class_name = self.seg_model1.names[int(cls_id)]
                    y, x = np.unravel_index(np.argmax(mask_resized), mask_resized.shape)
                    cv2.putText(frame, class_name, (x, y), cv2.FONT_HERSHEY_SIMPLEX,
                                0.5, (255, 255, 255), 1)
                    # SnapShoot the defect
                    self.autosave()
                    self.defect_class = class_name
                    #self.gps_coordinates = "kk"
                    if class_name is not None:
                        self.class_variable = class_name
                        # Add count and action
                        if class_name == "Defective Joint":
                            self.defective_joints_count += 1
                            self.defective_joints_label.config(text=f"Defective Joints: {self.defective_joints_count}")
                        elif class_name == "Missing Dogscrew":
                            self.missing_dogscrew_count += 1
                            self.missing_dogscrew_label.config(text=f"Missing Dogscrew: {self.missing_dogscrew_count}")
                        elif class_name == "Missing Fastener":
                            self.defective_fasteners_count += 1
                            self.defective_fasteners_label.config(
                                text=f"Defective Fasteners: {self.defective_fasteners_count}")

        if self.seg2_var.get():
            seg_results2 = self.seg_model2(frame, conf=self.conf_var.get(), verbose=False)[0]
            if seg_results2.masks is not None:
                for mask, cls_id in zip(seg_results2.masks.data.cpu().numpy(), seg_results2.boxes.cls.cpu().numpy()):
                    mask_resized = cv2.resize(mask, (frame.shape[1], frame.shape[0]))
                    combined_mask = np.maximum(combined_mask, (mask_resized * 255).astype(np.uint8))

                    # Get class name and draw label
                    class_name = self.seg_model2.names[int(cls_id)]
                    y, x = np.unravel_index(np.argmax(mask_resized), mask_resized.shape)
                    cv2.putText(frame, class_name, (x, y), cv2.FONT_HERSHEY_SIMPLEX,
                                0.5, (255, 255, 255), 1)
                    # SnapShoot the defect
                    self.autosave()
                    self.defect_class = class_name
                    #self.gps_coordinates = "kk"
                    if class_name is not None:
                        self.class_variable = class_name
                        # Add count
                        if class_name == "defective-rail":
                            self.defective_rail_count += 1
                            self.defective_rail_label.config(text=f"Defective Rails: {self.defective_rail_count}")

        # Apply color overlay for segmentation masks
        colored_mask = cv2.applyColorMap(combined_mask, cv2.COLORMAP_JET)
        overlay = cv2.addWeighted(frame, 0.7, colored_mask, 0.3, 0)

        # Process OBB detections if selected
        if self.obb_var.get():
            obb_results = self.obb_model(frame, conf=self.conf_var.get(), verbose=False)[0]
            if obb_results.obb is not None:
                for box, cls_id in zip(obb_results.obb.xyxyxyxy.cpu().numpy(), obb_results.obb.cls.cpu().numpy()):
                    pts = box.reshape(-1, 2).astype(np.int32)
                    cv2.polylines(overlay, [pts], isClosed=True, color=(0, 255, 255), thickness=2)

                    # Get class name and draw label
                    class_name = self.obb_model.names[int(cls_id)]
                    cv2.putText(overlay, class_name, (pts[0][0], pts[0][1] - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)
                    # SnapShoot the defect
                    self.autosave()
                    self.defect_class = class_name
                    #self.gps_coordinates = "kk"
                    if class_name is not None:
                        self.class_variable = class_name
                        # Add count
                        if class_name == "Fair Conrete Sleeper":
                            self.fair_sleepers_count += 1
                            self.fair_sleepers_label.config(text=f"Fair Sleepers: {self.fair_sleepers_count}")
                        elif class_name == "Good Concrete Sleeper":
                            self.good_sleepers_count += 1
                            self.good_sleepers_label.config(text=f"Good Sleepers: {self.poor_sleepers_count}")
                        elif class_name == "Poor Concrete Sleeper":
                            self.poor_sleepers_count += 1
                            self.poor_sleepers_label.config(text=f"Poor Sleepers: {self.poor_sleepers_count}")

        return overlay

    def show_processed_frame(self, frame):
        self.current_frame = frame
        self.update_canvas(frame)

    def update_canvas(self, frame):
        # Resize image to fit canvas while maintaining aspect ratio
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()

        if canvas_width <= 1 or canvas_height <= 1:
            canvas_width = 800
            canvas_height = 600

        img = Image.fromarray(frame)
        img.thumbnail((canvas_width, canvas_height))

        self.photo = ImageTk.PhotoImage(image=img)
        self.canvas.create_image(canvas_width // 2, canvas_height // 2, image=self.photo, anchor=tk.CENTER)

    def ultrasonic_window(self):
        '''for widget in self.winfo_children():
            widget.grid_remove() #destroy()
        self.master.withdraw()  # Hide the window
        #self.master.deiconify()  # Show it again
        UltrasonicInspectionGUI(self.master)'''

        # Clear current widgets
        for widget in self.master.winfo_children():
            widget.destroy()  # Use destroy instead of grid_remove for complete cleanup

        # Create new ultrasonic inspection GUI
        ultrasonic_gui = UltrasonicInspectionGUI(self.master)

        # Make sure the window is visible
        self.master.deiconify()  # Ensure window is shown
        self.master.state('zoomed')  # Maximize window

def center_window(width, height):
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
class WelcomeWindow(tk.Frame):
    def __init__(self, master):
        super().__init__()
        self.master = master
        self.master.title("Welcome")
        center_window(380, 250)
        self.master.resizable(False, False)
        self.master.configure(bg="#326273")

        self.frame = tk.Frame(self)
        self.frame.pack()
        self.frame.configure(bg="#326273")

        ###login infor
        self.login_frame = tkinter.LabelFrame(self.frame, text="Login Information", font="arial13", bg="#326273", fg="#fff")
        self.login_frame.grid(row=0, column=0, padx=20, pady=10)

        username_label = tkinter.Label(self.login_frame, text="Username", font="arial13", bg="#326273", fg="#fff")
        username_label.grid(row=0, column=0)
        self.username_entry = tkinter.Entry(self.login_frame, width=15, bd=1.5, font=20)
        self.username_entry.grid(row=0, column=1)

        password_label = tkinter.Label(self.login_frame, text="Password", font="arial13", bg="#326273", fg="#fff")
        password_label.grid(row=1, column=0)
        self.password_entry = tkinter.Entry(self.login_frame, show="*", width=15, bd=1.5, font=20)
        self.password_entry.grid(row=1, column=1)

        for widget in self.login_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Control buttons

        login_button = tkinter.Button(self.frame, text="Login", command=self.login, font="ariel 13",
                                      bg="#326273", fg="white", width=15, height=1)
        login_button.grid(row=1, column=0, sticky="nws", padx=20, pady=10)

        sign_up_button = tkinter.Button(self.frame, text="Sign Up", command=self.signup, font="ariel 13",
                                        bg="#326273", fg="white", width=15, height=1)
        sign_up_button.grid(row=1, column=0, sticky="nes", padx=20, pady=10)

        exit1_button = tkinter.Button(self.frame, text="Exit", command=self.close, font="ariel 13",
                                      bg="#326273", fg="white", width=15, height=1)
        exit1_button.grid(row=2, column=0, sticky="news", padx=20, pady=10)


        # login_button = tkinter.Button(self, text="Login", width=10, command=self.on_login)
        # login_button.pack(padx=20, pady=(20, 10))
        #
        # register_button = tkinter.Button(self, text="Register", width=10, command=self.on_register)
        # register_button.pack(pady=10)


        self.pack()

    def login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()

       # path = "C:/Users/ASUS/Desktop/Yolo/data.xlsx"
        path = "C:/Users/ASUS/Desktop/Level 4 Project/Software/code/program/MachineVision/data.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)
        for value_tuple in list_values[1:]:
            # print(value_tuple)
            if value_tuple[6] == username and value_tuple[7] == password:
                # print(username + " " + password)
                messagebox.showinfo("Login Successful", "Login Successful")
                for widget in self.winfo_children():
                    widget.destroy()
                self.destroy()
                YOLOv8_GUI(self.master)
            else:
                messagebox.showerror("Error", "Incorrect login details")
                break


    def signup(self):
        for widget in self.winfo_children():
            widget.destroy()
        self.destroy()
        SignUpWindow(self.master)

    def close(self):
        for widget in self.winfo_children():
            widget.destroy()
        self.destroy()
        root.destroy()
        #WelcomeWindow(self.master)

class SignUpWindow(tk.Frame):
    def __init__(self, master):
        super().__init__()
        self.master = master
        self.master.title("Sign Up")
        self.master.resizable(False, False)
        center_window(660, 560)
        self.configure(bg="#326273")
        self.pack()

        self.frame = tkinter.Frame(self)
        self.frame.pack()
        self.frame.configure(bg="#326273")

        ####Saving User Information
        self.user_info_frame = tkinter.LabelFrame(self.frame, text="User Information", font="arial13", bg="#326273", fg="#fff")
        self.user_info_frame.grid(row=0, column=0, padx=20, pady=10)

        first_name_label = tkinter.Label(self.user_info_frame, text="First Name", font="arial13", bg="#326273", fg="#fff")
        first_name_label.grid(row=0, column=0)
        last_name_label = tkinter.Label(self.user_info_frame, text="Last Name", font="arial13", bg="#326273", fg="#fff")
        last_name_label.grid(row=0, column=1)

        self.first_name_entry = tkinter.Entry(self.user_info_frame, width=15, bd=1.5, font=20)
        self.first_name_entry.grid(row=1, column=0)
        self.last_name_entry = tkinter.Entry(self.user_info_frame, width=15, bd=1.5, font=20)
        self.last_name_entry.grid(row=1, column=1)

        title_label = tkinter.Label(self.user_info_frame, text="Title", font="arial13", bg="#326273", fg="#fff")
        title_label.grid(row=0, column=2)
        self.title_combobox = ttk.Combobox(self.user_info_frame, values=("Mr", "Mrs", "Ms", "Dr", "Eng", ""),
                                      state="r", width=15, font=20)
        self.title_combobox.grid(row=1, column=2)
        self.title_combobox.set("Mr")

        age_label = tkinter.Label(self.user_info_frame, text="Age", font="arial13", bg="#326273", fg="#fff")
        age_label.grid(row=2, column=0)
        self.age_spinbox = ttk.Spinbox(self.user_info_frame, from_=18, to=100, increment=1, width=15, font=20)
        self.age_spinbox.grid(row=3, column=0)

        regnumber_label = tkinter.Label(self.user_info_frame, text="Registration Number", font="arial13", bg="#326273",
                                        fg="#fff")
        regnumber_label.grid(row=2, column=1)
        self.regnumber_entry = tkinter.Entry(self.user_info_frame, width=15, bd=1.5, font=20)
        self.regnumber_entry.grid(row=3, column=1)

        position_label = tkinter.Label(self.user_info_frame, text="Position", font="arial13", bg="#326273", fg="#fff")
        position_label.grid(row=2, column=2)
        self.position_combobox = ttk.Combobox(self.user_info_frame, values=("Operator", "Supervisor", "Manager", "Maintenance"),
                                         state="r", width=15, font=20)
        self.position_combobox.grid(row=3, column=2)
        self.position_combobox.set("Operator")

        for widget in self.user_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Login Information
        self.login_frame = tkinter.LabelFrame(self.frame, text="Login Information", font="arial13", bg="#326273", fg="#fff")
        self.login_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)

        username_label = tkinter.Label(self.login_frame, text="Username", font="arial13", bg="#326273", fg="#fff")
        username_label.grid(row=1, column=0)
        self.username_entry = tkinter.Entry(self.login_frame, width=15, bd=1.5, font=20)
        self.username_entry.grid(row=2, column=0)

        password_label = tkinter.Label(self.login_frame, text="Password", font="arial13", bg="#326273", fg="#fff")
        password_label.grid(row=1, column=1)
        self.password_entry = tkinter.Entry(self.login_frame, show="*", width=15, bd=1.5, font=20)
        self.password_entry.grid(row=2, column=1)

        conf_password_label = tkinter.Label(self.login_frame, text="Confirm Password", font="arial13", bg="#326273",
                                            fg="#fff")
        conf_password_label.grid(row=1, column=2)
        self.conf_password_entry = tkinter.Entry(self.login_frame, show="*", width=15, bd=1.5, font=20)
        self.conf_password_entry.grid(row=2, column=2)

        for widget in self.login_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        # Accept terms
        self.terms_frame = tkinter.LabelFrame(self.frame, text="Terms & Conditions", font="arial13", bg="#326273", fg="#fff")
        self.terms_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

        self.accept_var = tkinter.StringVar(value="Not Accepted")
        self.terms_check = tkinter.Checkbutton(self.terms_frame, text="I accept the terms and conditions.",
                                          variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted",
                                          font="arial13", bg="#326273")
        self.terms_check.grid(row=0, column=0)

        # Button
        button1 = tkinter.Button(self.frame, text="Save Data", command=self.save_data, font="ariel 13",
                                 bg="#326273", fg="white", height=2)
        button1.grid(row=3, column=0, sticky="news", padx=20, pady=10)

        # Control buttons

        button2 = tkinter.Button(self.frame, text="Back", command=self.back, font="ariel 13",
                                 bg="#326273", fg="white", width=30, height=2)
        button2.grid(row=4, column=0, sticky="nes", padx=20, pady=10)

        button3 = tkinter.Button(self.frame, text="Exit", command=self.close_sign_up, font="ariel 13",
                                 bg="#326273", fg="white", width=30, height=2)
        button3.grid(row=4, column=0, sticky="nws", padx=20, pady=10)


        self.pack()

    def back(self):
        for widget in self.winfo_children():
            widget.destroy()
        self.destroy()
        WelcomeWindow(self.master)

    def save_data(self):
        accepted = self.accept_var.get()

        if accepted == "Accepted":
            # User info
            firstname = self.first_name_entry.get()
            lastname = self.last_name_entry.get()
            title = self.title_combobox.get()
            age = self.age_spinbox.get()
            reg = self.regnumber_entry.get()
            position = self.position_combobox.get()

            if firstname and lastname and title and age and reg and position:

                # Login Infor
                username = self.username_entry.get()
                password = self.password_entry.get()
                confirmed = self.conf_password_entry.get()

                if password == confirmed:
                    #messagebox.showinfo("Confirmed Password")

                    filepath = "C:/Users/ASUS/Desktop/Level 4 Project/Software/code/program/MachineVision/data.xlsx"
                    #tkinter.messagebox.showinfo(title="Sign Up", message="Press save data again to save account"
                                                                         #" your data")
                    if not os.path.exists(filepath):
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active
                        heading = ["First Name", "Last Name", "Title", "Age", "Registration Number", "Position",
                                   "Username", "Password"]
                        sheet.append(heading)
                        workbook.save(filepath)
                    workbook = openpyxl.load_workbook(filepath)
                    sheet = workbook.active
                    sheet.append([firstname, lastname, title, age, reg, position, username, password])
                    workbook.save(filepath)
                    tkinter.messagebox.showinfo(title="Account Created", message="Your account was successfully "
                                                                 "created! /n You can now login to your account.")

                else:
                    tkinter.messagebox.showwarning(title ="Sign Up Error", message = "Password does not match")

            else:
                tkinter.messagebox.showwarning(title="Error", message="All user information fields are required.")
        else:
            tkinter.messagebox.showwarning(title="Error", message="You have not accepted the terms")


    def close_sign_up(self):
        for widget in self.winfo_children():
            widget.destroy()
        self.destroy()
        root.destroy()
        #WelcomeWindow(self.master)
class UltrasonicInspectionGUI(tk.Frame):
    def __init__(self, master):
        super().__init__(master)  # Parent is the main window
        self.master = master
        self.master.title("Ultrasonic Inspection System")
       # self.master.geometry("1200x800")
        self.grid(sticky="nsew")

        # Configure the main frame to expand
        #self.grid(row=0, column=0, sticky="nsew")

        # Configure grid weights for the root window
        self.master.grid_rowconfigure(0, weight=1)
        self.master.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

                # Make the window maximize (works on Windows/Linux)
        self.master.state('zoomed')  # or self.master.attributes('-fullscreen', True)

        # System status
        self.scanning = False
        self.connected = False
        # Serial Communication with Arduino
        self.port = None
        self.data = []
        # Initialize database
        self.db = UTImageDatabase()

        self.create_widgets()
    def create_widgets(self):
        # Create the container frame
        self.main_frame = ttk.Frame(self)
        self.main_frame.grid(row=0, column=0, sticky="nsew")

        # Configure grid weights for the main frame
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Create sub-frames
        control_frame = ttk.LabelFrame(self.main_frame, text="Control Panel", padding=10)
        display_frame = ttk.LabelFrame(self.main_frame, text="Scan Display", padding=10)
        config_frame = ttk.LabelFrame(self.main_frame, text="Configuration", padding=10)

        # Grid layout for sub-frames
        control_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        display_frame.grid(row=0, column=1, rowspan=2, padx=5, pady=5, sticky="nsew")
        config_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        # Configure weights for main_frame's grid
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(1, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(1, weight=4)  # Give more space to display

        # Add widgets to control_frame (example)
        self.status_label = ttk.Label(control_frame, text="Status: Disconnected", foreground="red")
        self.status_label.pack(fill=tk.X, pady=5)
        '''
        # Create main frames
        control_frame = ttk.LabelFrame(self.master, text="Control Panel", padding=10)
        display_frame = ttk.LabelFrame(self.master, text="Scan Display", padding=10)
        config_frame = ttk.LabelFrame(self.master, text="Configuration", padding=10)

        control_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
        display_frame.grid(row=0, column=1, rowspan=2, padx=5, pady=5, sticky="nsew")
        config_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        # Control Panel
        self.status_label = ttk.Label(control_frame, text="Status: Disconnected", foreground="red")
        self.status_label.pack(pady=5)'''
        self.connect_btn = ttk.Button(control_frame, text="Connect System", command=self.connect_system)
        self.connect_btn.pack(fill=tk.X, pady=5)

        self.scan_btn = ttk.Button(control_frame, text="Start Scan", command=self.toggle_scan)#, state=tk.DISABLED)
        self.scan_btn.pack(fill=tk.X, pady=5)

        self.history_btn = ttk.Button(control_frame, text="View History", command=self.history)
        self.history_btn.pack(fill=tk.X, pady=5)

        self.emergency_btn = ttk.Button(control_frame, text="EMERGENCY STOP", style='Emergency.TButton',
                                        command=self.emergency_stop)
        self.emergency_btn.pack(fill=tk.X, pady=5)

        self.back_btn = ttk.Button(control_frame, text="Back", command=self.back)
        self.back_btn.pack(fill=tk.X, pady=5)

        self.exit_btn = ttk.Button(control_frame, text="Exit", command=self.exit)
        self.exit_btn.pack(fill=tk.X, pady=5)

        self.autosave_ut_var = tk.BooleanVar(value=True)
        self.chk_mode2 = ttk.Checkbutton(control_frame, text="AutoSave", variable=self.autosave_ut_var,
                                         style="TCheckbutton")
        self.chk_mode2.pack(fill=tk.X, pady=5)

        # Add widgets to display_frame (example)
        self.setup_scan_display(display_frame)

        # Add widgets to config_frame (example)
        self.create_config_panel(config_frame)

        # Bind window resize event
        self.master.bind("<Configure>", self.on_window_resize)

        '''# Scan Display
        self.create_scan_display(display_frame)

        # Configuration
        self.create_config_panel(config_frame)'''

        # Configure grid weights
        self.master.columnconfigure(1, weight=1)
        self.master.rowconfigure(0, weight=1)

        # Style configuration
        style = ttk.Style()
        style.configure('Emergency.TButton', foreground='white', background='red', font=('Helvetica', 12, 'bold'))

    def on_window_resize(self, event):
        """Optional: Handle window resize events"""
        pass
    def back(self):
        # Clear current widgets
        for widget in self.master.winfo_children():
            widget.destroy()  # Use destroy instead of grid_remove for complete cleanup

        YOLOv8_GUI(self.master)
    def exit(self):
        self.master.destroy()
        #self.destroy()
    def history(self):
        return None
    def create_scan_display(self, parent):
        # Tabbed display for different scan types
        notebook = ttk.Notebook(parent)

        # A-Scan Tab
        a_scan_frame = ttk.Frame(notebook)
        self.setup_a_scan(a_scan_frame)
        notebook.add(a_scan_frame, text="A-Scan")

        # B-Scan Tab
        b_scan_frame = ttk.Frame(notebook)
        self.setup_b_scan(b_scan_frame)
        notebook.add(b_scan_frame, text="B-Scan")

        # C-Scan Tab
        c_scan_frame = ttk.Frame(notebook)
        self.setup_c_scan(c_scan_frame)
        notebook.add(c_scan_frame, text="C-Scan")

        notebook.pack(expand=True, fill=tk.BOTH)

    def setup_scan_display(self, parent):
        # Your existing scan display setup
        notebook = ttk.Notebook(parent)

        # A-Scan Tab
        a_scan_frame = ttk.Frame(notebook)
        self.setup_a_scan(a_scan_frame)
        notebook.add(a_scan_frame, text="A-Scan")

        # Pack notebook to fill parent
        notebook.pack(expand=True, fill=tk.BOTH)

    def setup_a_scan(self, frame):
        fig = Figure(figsize=(6, 4), dpi=100)
        self.a_scan_plot = fig.add_subplot(111)
        self.a_scan_plot.set_title("A-Scan Display")
        self.a_scan_plot.set_xlabel("Time (μs)")
        self.a_scan_plot.set_ylabel("Amplitude")

        self.a_scan_canvas = FigureCanvasTkAgg(fig, master=frame)
        self.a_scan_canvas.draw()
        self.a_scan_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Initial empty plot
        self.update_a_scan([], [])

    def setup_b_scan(self, frame):
        fig = Figure(figsize=(6, 4), dpi=100)
        self.b_scan_plot = fig.add_subplot(111)
        self.b_scan_plot.set_title("B-Scan Display")
        self.b_scan_plot.set_xlabel("Position")
        self.b_scan_plot.set_ylabel("Depth")

        self.b_scan_canvas = FigureCanvasTkAgg(fig, master=frame)
        self.b_scan_canvas.draw()
        self.b_scan_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def setup_c_scan(self, frame):
        fig = Figure(figsize=(6, 4), dpi=100)
        self.c_scan_plot = fig.add_subplot(111)
        self.c_scan_plot.set_title("C-Scan Display")
        self.c_scan_plot.set_xlabel("X Position")
        self.c_scan_plot.set_ylabel("Y Position")

        self.c_scan_canvas = FigureCanvasTkAgg(fig, master=frame)
        self.c_scan_canvas.draw()
        self.c_scan_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def create_config_panel(self, parent):
        # Notebook for different configuration sections
        notebook = ttk.Notebook(parent)

        # Transducer Settings
        transducer_frame = ttk.Frame(notebook)
        self.create_transducer_panel(transducer_frame)
        notebook.add(transducer_frame, text="Transducer")

        # Acquisition Settings
        acquisition_frame = ttk.Frame(notebook)
        self.create_acquisition_panel(acquisition_frame)
        notebook.add(acquisition_frame, text="Acquisition")

        # Motion Settings
        motion_frame = ttk.Frame(notebook)
        self.create_motion_panel(motion_frame)
        notebook.add(motion_frame, text="Motion")

        notebook.pack(expand=True, fill=tk.BOTH)

    def create_transducer_panel(self, frame):
        ttk.Label(frame, text="Frequency (MHz):").grid(row=0, column=0, sticky="w", pady=2)
        self.freq_var = tk.StringVar(value="5.0")
        ttk.Entry(frame, textvariable=self.freq_var).grid(row=0, column=1, sticky="ew", pady=2)

        ttk.Label(frame, text="Pulse Voltage (V):").grid(row=1, column=0, sticky="w", pady=2)
        self.voltage_var = tk.StringVar(value="100")
        ttk.Entry(frame, textvariable=self.voltage_var).grid(row=1, column=1, sticky="ew", pady=2)

        ttk.Label(frame, text="Probe Type:").grid(row=2, column=0, sticky="w", pady=2)
        self.probe_var = tk.StringVar()
        ttk.Combobox(frame, textvariable=self.probe_var,
                     values=["Immersion", "Contact", "Dual Element", "Phased Array"]).grid(row=2, column=1, sticky="ew",
                                                                                           pady=2)
    def create_acquisition_panel(self, frame):
        ttk.Label(frame, text="Gain (dB):").grid(row=0, column=0, sticky="w", pady=2)
        self.gain_var = tk.DoubleVar(value=40.0)
        ttk.Scale(frame, from_=0, to=100, variable=self.gain_var,
                  command=lambda v: self.gain_value.set(f"{float(v):.1f} dB")).grid(row=0, column=1, sticky="ew",
                                                                                    pady=2)
        self.gain_value = tk.StringVar()
        ttk.Label(frame, textvariable=self.gain_value).grid(row=0, column=2, sticky="w", pady=2)

        ttk.Label(frame, text="Range (μs):").grid(row=1, column=0, sticky="w", pady=2)
        self.range_var = tk.StringVar(value="50")
        ttk.Entry(frame, textvariable=self.range_var).grid(row=1, column=1, sticky="ew", pady=2)

        ttk.Label(frame, text="Filter:").grid(row=2, column=0, sticky="w", pady=2)
        self.filter_var = tk.StringVar()
        ttk.Combobox(frame, textvariable=self.filter_var,
                     values=["None", "Low Pass", "High Pass", "Band Pass"]).grid(row=2, column=1, sticky="ew", pady=2)

    def create_motion_panel(self, frame):
        ttk.Label(frame, text="Scan Area (mm):").grid(row=0, column=0, sticky="w", pady=2)

        ttk.Label(frame, text="X:").grid(row=1, column=0, sticky="w", pady=2)
        self.x_size_var = tk.StringVar(value="100")
        ttk.Entry(frame, textvariable=self.x_size_var).grid(row=1, column=1, sticky="ew", pady=2)

        ttk.Label(frame, text="Y:").grid(row=2, column=0, sticky="w", pady=2)
        self.y_size_var = tk.StringVar(value="100")
        ttk.Entry(frame, textvariable=self.y_size_var).grid(row=2, column=1, sticky="ew", pady=2)

        ttk.Label(frame, text="Step Size (mm):").grid(row=3, column=0, sticky="w", pady=2)
        self.step_var = tk.StringVar(value="1.0")
        ttk.Entry(frame, textvariable=self.step_var).grid(row=3, column=1, sticky="ew", pady=2)

        ttk.Label(frame, text="Speed (mm/s):").grid(row=4, column=0, sticky="w", pady=2)
        self.speed_var = tk.StringVar(value="10")
        ttk.Entry(frame, textvariable=self.speed_var).grid(row=4, column=1, sticky="ew", pady=2)

    def toggle_scan(self):
        if not self.scanning:
            self.scanning = True
            self.scan_btn.config(text="Stop Scan")
            self.start_scanning()
        else:
            self.scanning = False
            self.scan_btn.config(text="Start Scan")

    def start_scanning(self):
        if self.scanning:
            # Convert ADC values to voltage (12-bit, 3.3V reference)
            te = np.linspace(0, float(self.range_var.get()), 500)
            amplitude = np.sin(2 * np.pi * float(self.freq_var.get()) * te) * (float(self.gain_var.get()) / 100)

            adc_data = np.array([amplitude], dtype=np.uint16)  # From Arduino
            voltage = adc_data * (3.3 / 4095)

            # # Simulate raw ADC data (5MHz signal + noise)
            # fs = 50e6  # 50 MSPS
            # time = np.arange(0, 1000) / fs
            # raw_signal = (np.sin(2 * np.pi * 5e6 * time) + 0.3 * np.random.randn(len(time))  # 5MHz + noise
            #
            # Time axis
            fs = 50e6  # 50 MSPS
            time = np.arange(len(voltage)) / fs * 1e6  # μs
            # amplitude = None

            """Plot frequency response of the filter."""
            if self.filter_var == 'Band Pass':
                filtered_voltage = self.ultrasonic_bandpass_filter(
                    voltage,
                    sample_rate=50e6,
                    center_freq=5e6,
                    bandwidth=2e6,
                    order=4,
                    plot_response=False)
                amplitude = filtered_voltage

            elif self.filter_var == 'Low Pass':
                filtered_voltage = self.lowpass(
                    voltage,
                    sample_rate=50e6,
                    center_freq=5e6,
                    bandwidth=2e6,
                    cutoff_freq=25e6,
                    order=4
                )
                amplitude = filtered_voltage


            elif self.filter_var == 'High Pass':
                filtered_voltage = self.highpass(
                    voltage,
                    sample_rate=50e6,
                    center_freq=5e6,
                    bandwidth=2e6,
                    cutoff_freq=100e6,
                    order=2
                )
                amplitude = filtered_voltage

            elif self.filter_var == 'None':
                amplitude = voltage

            # Detect peaks
            peaks, tof, envelope = self.detect_ultrasonic_peaks(amplitude)

            # Update plots
            self.update_a_scan(time, amplitude)
            self.update_b_scan()
            self.update_c_scan()

            # Schedule next update
            self.master.after(100, self.start_scanning)

    def update_a_scan(self, time, amplitude):
        self.a_scan_plot.clear()
        ascan_signal = amplitude
        num_points = len(amplitude)
        time = np.linspace(0, 1, num_points)  # 500 points from 0 to 1 second

        # Detect peaks (using TDC7201-like thresholding)
        peaks, _ = find_peaks(ascan_signal, height=0.2, distance=100)

        if len(time) > 0:
            print("Time array shape:", time)
            print("Signal array shape:", ascan_signal)
            print("Peaks:", peaks)
            print("amplitude:", amplitude)

            # Filter valid peaks
            valid_peaks = [p for p in peaks if p < len(ascan_signal)]

            self.a_scan_plot.plot(time, amplitude, 'b-')
            # Plot only if we found valid peaks
            if len(valid_peaks) > 0:
                self.a_scan_plot.scatter(
                    time[valid_peaks] * 1e6,
                    ascan_signal[valid_peaks],
                    color='red',
                    label="Detected Peaks"
                )
            else:
                print("Warning: No valid peaks found!")
            # a_scan_plot.scatter(time[peaks] * 1e6, ascan_signal[peaks], color='red', label="Detected Peaks")
            self.a_scan_plot.set_title("A-Scan Display")
            self.a_scan_plot.set_xlabel("Time (μs)")
            self.a_scan_plot.set_ylabel("Amplitude")
            self.a_scan_plot.grid(True)
        self.a_scan_canvas.draw()

        # Take snapshots
        if self.autosave_ut_var.get() and peaks is not None:
            # take a snapshot
            # Save the figure with professional settings
            save_dir = "saved_plots"
            os.makedirs(save_dir, exist_ok=True)  # Create directory if needed
            gps_coordinates = self.getgps()
            filename = f"{save_dir}/{gps_coordinates}/ascan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"

            try:
                self.a_scan_plot.savefig(
                    filename,
                    dpi=300,  # High resolution
                    bbox_inches='tight',  # Remove extra whitespace
                    facecolor='white',  # Background color
                    format='png'  # Can also use 'pdf', 'svg', etc.
                )
                print(f"Successfully saved to {filename}")
            except Exception as e:
                print(f"Error saving figure: {e}")
            # finally:
            #     plt.close(self.a_scan_plot)  # Always close the figure to free memory

    def getgps(self):
        return None

    def update_b_scan(self):
        # Simulate B-scan data
        x = np.linspace(0, float(self.x_size_var.get()), 50)
        y = np.linspace(0, float(self.y_size_var.get()), 50)
        X, Y = np.meshgrid(x, y)
        Z = np.sin(X) * np.cos(Y) * float(self.gain_var.get()) / 100

        self.b_scan_plot.clear()
        self.b_scan_plot.imshow(Z, cmap='hot',
                                extent=[0, float(self.x_size_var.get()), 0, float(self.y_size_var.get())])
        self.b_scan_plot.set_title("B-Scan Display")
        self.b_scan_plot.set_xlabel("Position (mm)")
        self.b_scan_plot.set_ylabel("Depth (mm)")
        self.b_scan_canvas.draw()

    def update_c_scan(self):
        # Simulate C-scan data
        x = np.linspace(0, float(self.x_size_var.get()), 50)
        y = np.linspace(0, float(self.y_size_var.get()), 50)
        X, Y = np.meshgrid(x, y)
        Z = np.sin(0.5 * X) * np.cos(0.5 * Y) * float(self.gain_var.get()) / 100

        self.c_scan_plot.clear()
        contour = self.c_scan_plot.contourf(X, Y, Z, levels=20, cmap='viridis')
        self.c_scan_plot.set_title("C-Scan Display")
        self.c_scan_plot.set_xlabel("X Position (mm)")
        self.c_scan_plot.set_ylabel("Y Position (mm)")
        fig = self.c_scan_plot.get_figure()
        fig.colorbar(contour, ax=self.c_scan_plot)
        self.c_scan_canvas.draw()

    def emergency_stop(self):
        self.scanning = False
        self.scan_btn.config(text="Start Scan")
        messagebox.showerror("Emergency Stop", "All systems halted!")

    def find_arduino_port(self):
        """Identify Arduino port automatically"""
        ports = serial.tools.list_ports.comports()
        for port in ports:
            if 'Arduino' in port.description or 'USB Serial Device' in port.description:
                return port.device
        return None

    def connect_system(self):
        if not self.connected:
            try:
                # Auto-detect port
                self.port = self.find_arduino_port()

                if self.port is None:
                    messagebox.showerror("Serial Connection Error", "Arduino not found! Check your connection.")
                    raise Exception("Arduino not found! Check your connection.")

                # try:
                #     messagebox
                print(f"Connecting to {self.port}...")
                messagebox.showinfo("Serial Connection", f"Connecting to {self.port}...")
                self.ser = serial.Serial(self.port, 115200, timeout=1)
                time.sleep(2)  # Wait for Arduino to initialize

                # Simulate connection
                self.connected = True
                self.status_label.config(text="Status: Connected", foreground="green")
                self.scan_btn.config(state=tk.NORMAL)
                self.connect_btn.config(text="Disconnect System")
                messagebox.showinfo("System", "Ultrasonic system connected successfully")

                while True:
                    if self.ser.in_waiting:
                        data = self.ser.readline().decode('utf-8').strip()
                        print(f"Received: {data}")

                    time.sleep(1)

            except serial.SerialException as e:
                print(f"Serial error: {e}")
            except Exception as e:
                print(f"Error: {e}")
            finally:
                if 'ser' in locals() and self.ser.is_open:
                    self.ser.close()
                print("Port closed")
        else:
            self.connected = False
            self.status_label.config(text="Status: Disconnected", foreground="red")
            self.scan_btn.config(state=tk.DISABLED, text="Start Scan")
            self.connect_btn.config(text="Connect System")
            self.scanning = False
            messagebox.showinfo("System", "System disconnected")

    # Python parsing example
    def read_serial(self, num_samples):
        while True:
            self.ser.reset_input_buffer()
            line = self.ser.readline().decode().strip()
            if line == "BEGIN_ECHO_DATA":

                while True:
                    line = self.ser.readline().decode().strip()
                    if line == "END_ECHO_DATA":
                        return np.array(self.data, dtype=np.uint16)
                    self.data.append(int(line))

    def ultrasonic_bandpass_filter(
            self,
            input_signal: np.ndarray,
            sample_rate: float = 50e6,  # 50 MSPS (AD9271 default)
            center_freq: float = 5e6,  # 5 MHz transducer
            bandwidth: float = 2e6,  # ±1 MHz bandwidth
            order: int = 4,  # Filter order (higher = steeper roll-off)
            plot_response: bool = False) -> np.ndarray:
        """
        Applies a Butterworth bandpass filter to ultrasonic signals.

        Args:
            input_signal: Raw ADC data (1D numpy array)
            sample_rate: Sampling rate in Hz
            center_freq: Center frequency in Hz
            bandwidth: Total bandwidth in Hz
            order: Filter order (must be even)
            plot_response: If True, plots filter frequency response

        Returns:
            Filtered signal (same shape as input)
        """
        # Calculate cutoff frequencies
        nyquist = 0.5 * sample_rate
        low = (center_freq - bandwidth / 2) / nyquist
        high = (center_freq + bandwidth / 2) / nyquist

        # Design Butterworth bandpass filter
        b, a = signal.butter(order, [low, high], btype='bandpass')

        ### Optionally plot frequency response
        if plot_response:
            w, h = signal.freqz(b, a, fs=sample_rate)
            plt.figure()
            plt.plot(w, 20 * np.log10(abs(h)))
            plt.title(f"Bandpass Filter Response ({center_freq / 1e6}MHz ± {bandwidth / 2 / 1e6}MHz)")
            plt.xlabel('Frequency (Hz)')
            plt.ylabel('Gain (dB)')
            plt.grid()
            plt.show()

        # Apply zero-phase filtering (preserves time alignment)
        return signal.filtfilt(b, a, input_signal)

    def lowpass(
            self,
            input_signal: np.ndarray,
            sample_rate: float = 50e6,  # 50 MSPS (AD9271 default)
            center_freq: float = 5e6,  # 5 MHz transducer
            bandwidth: float = 2e6,  # ±1 MHz bandwidth
            cutoff_freq: float = 25e6,
            order: int = 4  # Filter order (higher = steeper roll-off)
    ) -> np.ndarray:
        """
        Anti-aliasing low-pass filter.

        Args:
            cutoff_freq: -3dB cutoff frequency (≤0.5*sample_rate)
        """
        # Calculate cutoff frequencies
        nyquist = 0.5 * sample_rate
        low = (center_freq - bandwidth / 2) / nyquist
        high = (center_freq + bandwidth / 2) / nyquist

        normalized_cutoff = cutoff_freq / nyquist
        b, a = signal.butter(order, normalized_cutoff, btype='low')
        return signal.filtfilt(b, a, input_signal)

    def highpass(
            self,
            input_signal: np.ndarray,
            sample_rate: float = 50e6,  # 50 MSPS (AD9271 default)
            center_freq: float = 5e6,  # 5 MHz transducer
            bandwidth: float = 2e6,  # ±1 MHz bandwidth
            cutoff_freq: float = 100e6,
            order: int = 2  # Filter order (higher = steeper roll-off)
    ) -> np.ndarray:

        """
                DC-blocking high-pass filter.

                Args:
                    cutoff_freq: Start frequency for passband
                """
        # Calculate cutoff frequencies
        nyquist = 0.5 * sample_rate
        low = (center_freq - bandwidth / 2) / nyquist
        high = (center_freq + bandwidth / 2) / nyquist

        normalized_cutoff = cutoff_freq / nyquist
        b, a = signal.butter(order, normalized_cutoff, btype='high')
        return signal.filtfilt(b, a, input_signal)

    def calculate_tof(self, envelope):
        threshold = 0.5 * np.max(envelope)
        crossing = np.where(envelope > threshold)[0][0]
        return crossing * (1 / 50e6)  # 50MSPS → time in seconds

    def detect_ultrasonic_peaks(self, voltage_signal, sample_rate=50e6, min_height=0.1, min_distance=10e-6):
        """
        Detect peaks in ultrasonic echoes with time-of-flight calculation.

        Args:
            voltage_signal: Filtered signal (in Volts)
            sample_rate: Sampling rate in Hz (50MSPS default)
            min_height: Minimum peak height (fraction of max amplitude)
            min_distance: Minimum time between peaks (in seconds)

        Returns:
            peaks: Indices of detected peaks
            tof: Time-of-flight for first peak (in seconds)
        """
        # 1. Normalize signal to [0,1] range
        norm_signal = voltage_signal / np.max(voltage_signal)

        # 2. Hilbert transform for envelope detection
        analytic_signal = signal.hilbert(norm_signal)
        envelope = np.abs(analytic_signal)

        # 3. Find peaks in the envelope
        min_samples = int(min_distance * sample_rate)  # Convert time to samples
        peaks, _ = signal.find_peaks(
            envelope,
            height=min_height,
            distance=min_samples,
            prominence=0.3  # Ignores small fluctuations
        )

        # 4. Calculate time-of-flight (for first echo)
        tof = peaks[0] / sample_rate if len(peaks) > 0 else None

        return peaks, tof, envelope

if __name__ == "__main__":
    root = tk.Tk()
    root.eval('tk::PlaceWindow . center')
    app = WelcomeWindow(root)
    #root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()